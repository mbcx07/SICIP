#!/usr/bin/env python3
"""
SICIP - Seed de Módulos desde CAPTURA_DE_CORRESPONDENCIA_2019.xlsm
Sube datos a Firestore para poblar las nuevas colecciones.
"""
import openpyxl
import json
import requests
import sys
import time
from datetime import datetime

XLSX = '/home/maxter/.openclaw/workspace/SICIP/CAPTURA_DE_CORRESPONDENCIA_2019.xlsm'
FIREBASE_API_KEY = "AIzaSyCBs_vl7IZ98Cr-Hs3VdVuDJyKPZetfOW8"
PROJECT_ID = "sicip-bcs"

# ─── Auth ──────────────────────────────────────────────────────
def get_token():
    r = requests.post(
        f"https://identitytoolkit.googleapis.com/v1/accounts:signInWithPassword?key={FIREBASE_API_KEY}",
        json={"email":"moises.beltran@imss.gob.mx","password":"LuMo221407","returnSecureToken":True}
    )
    return r.json().get("idToken", "")

TOKEN = get_token()
HEADERS = {"Authorization": f"Bearer {TOKEN}", "Content-Type": "application/json"}

# ─── Helpers ───────────────────────────────────────────────────
def firestore_post(collection, body):
    url = f"https://firestore.googleapis.com/v1/projects/{PROJECT_ID}/databases/(default)/documents/{collection}"
    r = requests.post(url, headers=HEADERS, json=body)
    if r.status_code != 200:
        print(f"    ❌ {r.status_code}: {r.text[:150]}")
    return r.status_code == 200

def firestore_set(collection, doc_id, body):
    url = f"https://firestore.googleapis.com/v1/projects/{PROJECT_ID}/databases/(default)/documents/{collection}/{doc_id}"
    r = requests.patch(url, headers=HEADERS, json=body)
    return r.status_code == 200

def make_fields(data):
    """Convert Python dict to Firestore Fields format."""
    fields = {}
    for k, v in data.items():
        if v is None:
            continue
        if isinstance(v, bool):
            fields[k] = {"booleanValue": v}
        elif isinstance(v, int):
            fields[k] = {"integerValue": str(v)}
        elif isinstance(v, float):
            fields[k] = {"doubleValue": v}
        elif isinstance(v, dict):
            fields[k] = {"mapValue": {"fields": make_fields(v)}}
        elif isinstance(v, list):
            fields[k] = {"arrayValue": {"values": [{"stringValue": str(x)} for x in v if x is not None]}}
        else:
            fields[k] = {"stringValue": str(v)}
    return fields

UPLOAD_LIMIT = None  # None = all; set to small number to test

# ══════════════════════════════════════════════════════════════
print("=" * 60)
print("SICIP SEED — Desde Excel a Firestore")
print("=" * 60)
print(f"Autenticado: {'SÍ' if TOKEN else 'NO'}")

wb = openpyxl.load_workbook(XLSX, keep_vba=False, data_only=True)

# ─── 1. DOCUMENTOS ──────────────────────────────────────────
print(f"\n{'─'*60}\n📁 1. DOCUMENTOS ({wb["DOCUMENTOS"].max_row} tipos)")
ws = wb['DOCUMENTOS']
count = 0
for r in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
    name = str(r[0]).strip() if r[0] else ''
    if not name or name == 'OTRO':
        continue
    cat = 'OTRO'
    for kw, c in [('INCAPACIDAD','INCAPACIDAD'),('LICENCIA','LICENCIA'),('COMISION','LICENCIA'),
                  ('CONTRATO','CONTRATO'),('NOMINA','NOMINA'),('VACACIONES','PRESTACIONES'),
                  ('GUARDIA','PRESTACIONES'),('TIEMPO','PRESTACIONES'),('NIVELACION','CONTRATO'),
                  ('BECAS','PERSONAL'),('INDUCCION','PERSONAL'),('BALANCE','NOMINA')]:
        if kw in name.upper():
            cat = c
            break
    ok = firestore_post('tipos_documentales', {"fields": make_fields({
        "clave": f"DOC-{count+1:02d}",
        "nombre": name,
        "categoria": cat,
        "descripcion": "",
        "requiereFirma": True,
        "requiereSello": True,
        "activo": True,
    })})
    if ok:
        count += 1
    if UPLOAD_LIMIT and count >= UPLOAD_LIMIT: break
print(f"  ✅ {count} tipos documentales en Firestore")

# ─── 2. INDICADORES ANP → indicadores_rh ────────────────────
print(f"\n{'─'*60}\n📈 2. INDICADORES ANP")
ws = wb['INDICADORES ANP']
print(f"  Filas: {ws.max_row}, Columnas: {ws.max_column}")
# Print headers to understand structure
for r in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
    vals = {f"col{i}": str(v)[:50] for i, v in enumerate(r) if v is not None}
    if vals:
        print(f"  {vals}")

# ─── 3. Sueldos → categorias_salariales ─────────────────────
print(f"\n{'─'*60}\n💰 3. CATEGORÍAS SALARIALES")
ws = wb['Sueldos']
data = list(ws.iter_rows(min_row=2, max_row=min(UPLOAD_LIMIT or ws.max_row, ws.max_row), values_only=True))
count = 0
for row in data:
    if not row[0]: continue
    nombre = str(row[0]).strip()
    smi = float(str(row[1] or '0').replace('$','').replace(',','').replace(' ','')) or 0
    ok = firestore_post('categorias_salariales', {"fields": make_fields({
        "clave": f"CAT-{count+1:04d}",
        "nombre": nombre,
        "smi": smi,
        "salarioDiario": round(smi / 30, 2),
        "salarioMensual": smi,
        "factorIntegracion": 1.0,
        "tipoContrato": "BASE",
        "vigenciaInicio": "2024-01-01",
        "activo": True,
    })})
    if ok:
        count += 1
print(f"  ✅ {count} categorías salariales")

# ─── 4. CATEHOR → puestos ──────────────────────────────────
print(f"\n{'─'*60}\n🏷️ 4. PUESTOS (CATEHOR)")
ws = wb['CATEHOR']
data = list(ws.iter_rows(min_row=3, max_row=min(UPLOAD_LIMIT or ws.max_row, ws.max_row), values_only=True))
count = 0
for row in data:
    if not row[1]: continue  # CVE PUESTO
    clave_puesto = str(row[1]).strip()
    categoria = str(row[2] or '').strip()
    s_afil = str(row[3] or '0').replace('$','').replace(',','')
    s_contrato = str(row[4] or '0').replace('$','').replace(',','')
    horas = str(row[5] or '').strip()
    descripcion = str(row[8] or '').strip()
    horario_txt = str(row[7] or '').strip()

    try:
        s_afil_float = float(s_afil)
    except:
        s_afil_float = 0
    try:
        s_contrato_float = float(s_contrato)
    except:
        s_contrato_float = 0

    ok = firestore_post('puestos', {"fields": make_fields({
        "clave": clave_puesto,
        "nombre": categoria,
        "descripcion": descripcion,
        "tipo": "OTRO",
        "nivel": "",
        "horario": {"entrada": "08:00", "salida": "16:00", "descanso": horario_txt},
        "salarioAfil": round(s_afil_float, 2),
        "salarioContrato": round(s_contrato_float, 2),
        "activo": True,
    })})
    if ok:
        count += 1
print(f"  ✅ {count} puestos en Firestore")

# ─── 5. Calendario Laboral ──────────────────────────────────
print(f"\n{'─'*60}\n📅 5. CALENDARIO LABORAL")
count = 0
for year in [2024, 2025, 2026]:
    for m in range(1, 13):
        import calendar
        for d in range(1, calendar.monthrange(year, m)[1] + 1):
            dt = datetime(year, m, d)
            dia_sem = dt.weekday()  # 0=Monday
            dia_sem_dom = (dia_sem + 1) % 7  # 0=Sunday
            tipo = 'INHABIL' if dia_sem_dom in (0, 6) else 'LABORAL'
            fecha = dt.strftime('%Y-%m-%d')

            # Check if already exists
            ok = firestore_post('calendario_laboral', {"fields": make_fields({
                "fecha": fecha,
                "tipo": tipo,
                "descripcion": "Sábado" if dia_sem_dom == 6 else ("Domingo" if dia_sem_dom == 0 else "Día laboral"),
                "aplicaA": "TODOS",
                "año": year,
                "mes": m,
                "diaSemana": dia_sem_dom,
                "esFinde": dia_sem_dom in (0, 6),
            })})
            if ok:
                count += 1
            if UPLOAD_LIMIT and count >= UPLOAD_LIMIT: break
        if UPLOAD_LIMIT and count >= UPLOAD_LIMIT: break
    if UPLOAD_LIMIT and count >= UPLOAD_LIMIT: break
print(f"  ✅ {count} días de calendario en Firestore (2024-2026)")

print(f"\n{'='*60}")
print(f"🌱 SEED COMPLETADO")
print(f"{'='*60}")
